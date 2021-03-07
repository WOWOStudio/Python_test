from openpyxl import Workbook,load_workbook
'''
wb = Workbook()
sheet = wb.active
sheet.title = '考勤统计表'
data = [
    ['姓名','出勤天数','迟到次数'],
    ['张三',20,5],
    ['李四',22,0]
]
for row in data:
    sheet.append(row)
wb.save('考勤统计.xlsx')
'''
wb = load_workbook('考勤统计.xlsx')
sheet  = wb['考勤统计表']
print(wb.sheetnames)
print(sheet['A1'].value)
for row in sheet.rows:
    for cell in row:
        print(cell.value)
